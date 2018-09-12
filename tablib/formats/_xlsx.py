# -*- coding: utf-8 -*-

""" Tablib - XLSX Support.
"""

import sys

import openpyxl

if sys.version_info[0] > 2:
    from io import BytesIO
else:
    from cStringIO import StringIO as BytesIO

from tempfile import NamedTemporaryFile
import xlsxwriter
import tablib

# TODO: fix these:
Workbook = openpyxl.workbook.Workbook
ExcelWriter = openpyxl.writer.excel.ExcelWriter
get_column_letter = openpyxl.utils.get_column_letter

from tablib.compat import unicode


title = 'xlsx'
extensions = ('xlsx',)


def detect(stream):
    """Returns True if given stream is a readable excel file."""
    try:
        openpyxl.reader.excel.load_workbook(stream)
        return True
    except openpyxl.shared.exc.InvalidFileException:
        pass

def export_set(dataset, freeze_panes=True):
    """Returns XLSX representation of Dataset."""
    wb, temp_file, cell_formats = _new_workbook([dataset])
    # TODO: adapt this openpyxl code to xlsxwriter:
    ws = wb.add_worksheet()
    ws.title = dataset.title if dataset.title else 'Tablib Dataset'

    dset_sheet(dataset, ws, cell_formats)
    wb.close()

    stream = BytesIO()
    print temp_file
    with open(temp_file, 'rb') as f:
        stream.write(f.read())
    return stream.getvalue()


def _create_format(workbook, format_dict):
    """
    creating xlsxwriter format
    :param workbook: xlsxwriter object
    :param format_dict: dictionary with format parameters
    :return: xlsxwriter format object
    """
    cell_format = workbook.add_format()
    if format_dict.get('bg_color'):
        cell_format.set_bg_color(format_dict['bg_color'])
    if format_dict.get('font'):
        cell_format.set_font_name(format_dict['font'])
    if format_dict.get('font_color'):
        cell_format.set_font_color(format_dict['font_color'])
    if format_dict.get('font_size'):
        cell_format.set_font_size(format_dict['font_size'])
    if format_dict.get('bold'):
        cell_format.set_bold()
    if format_dict.get('italic'):
        cell_format.set_italic()
    if format_dict.get('aligment'):
        cell_format.set_align(format_dict['aligment'])
    if format_dict.get('border'):
        cell_format.set_border(format_dict['border'])
    return cell_format


def _format_dict_key(format_dict):
    if not format_dict:
        return None
    return tuple(sorted(format_dict.items()))


def _new_workbook(datasets):
    """
    creating xlsxwriter workbook
    :param datasets: dataset object
    :return: workbook xlsxwriter object, filename str, cell format xlsxwriter object
    """
    temp_file = NamedTemporaryFile()
    workbook = xlsxwriter.Workbook('{}.xlsx'.format(temp_file.name))
    cell_formats = {}
    for dataset in datasets:
        for format_dict in dataset.formats:
            cell_format = _create_format(workbook, format_dict)
            cell_formats[_format_dict_key(format_dict)] = cell_format
        if dataset.conditional_formats:
            for cond_format in dataset.conditional_formats:
                format = _create_format(workbook, cond_format['options']['format'])
                cond_format['options']['format'] = format
    return workbook, temp_file.name+'.xlsx', cell_formats


def export_book(databook, freeze_panes=True):
    """Returns XLSX representation of DataBook."""
    wb, temp_file, cell_formats = _new_workbook(databook)
    for i, dset in enumerate(databook._datasets):
        ws = wb.create_sheet()
        ws.title = dset.title if dset.title else 'Sheet%s' % (i)

        dset_sheet(dset, ws, cell_formats)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def import_set(dset, in_stream, headers=True):
    """Returns databook from XLS stream."""

    dset.wipe()

    xls_book = openpyxl.reader.excel.load_workbook(BytesIO(in_stream))
    sheet = xls_book.get_active_sheet()

    dset.title = sheet.title

    for i, row in enumerate(sheet.rows):
        row_vals = [c.value for c in row]
        if (i == 0) and (headers):
            dset.headers = row_vals
        else:
            dset.append(row_vals)


def import_book(dbook, in_stream, headers=True):
    """Returns databook from XLS stream."""

    dbook.wipe()

    xls_book = openpyxl.reader.excel.load_workbook(BytesIO(in_stream))

    for sheet in xls_book.worksheets:
        data = tablib.Dataset()
        data.title = sheet.title

        for i, row in enumerate(sheet.rows):
            row_vals = [c.value for c in row]
            if (i == 0) and (headers):
                data.headers = row_vals
            else:
                data.append(row_vals)

        dbook.add_sheet(data)


def dset_sheet(dataset, ws, cell_formats):
    """Completes given worksheet from given Dataset."""
    _package = dataset._package(dicts=False, formats=True)
    for i, sep in enumerate(dataset._separators):
        _offset = i
        _package.insert((sep[0] + _offset), (sep[1],))

    if dataset.dropdowns:
        for dropdown in dataset.dropdowns:
            ws.data_validation(**dropdown)

    if dataset.conditional_formats:
        for cond_format in dataset.conditional_formats:
            ws.conditional_format(**cond_format)

    for i, row in enumerate(_package):
        for j, (col, format) in enumerate(row):
            cell_format = cell_formats.get(_format_dict_key(format))
            ws.write(i, j, col, cell_format)



